# Здесь будут определены задачи Celery для обработки файлов
from celery import shared_task # Убираем Celery, если не создаем экземпляр app здесь
import time
import os
import zipfile # Пока не используем, можно закомментировать или оставить
# Важно импортировать модель ProcessingTask
# Но чтобы избежать циклических импортов, если tasks.py импортируется Django при старте,
# лучше делать импорт модели внутри самой задачи или использовать app.loader.get_model

# Если используется Django (как у нас), то приложение Celery уже должно быть настроено
# и можно использовать @shared_task. 
# Для доступа к моделям Django внутри задачи, Django должен быть инициализирован.
# Обычно это происходит автоматически, когда worker запускается с -A proj.celery

# Прямой импорт модели ProcessingTask
from .models import ProcessingTask
from django.conf import settings # Для доступа к MEDIA_ROOT
from pathlib import Path # Для удобной работы с путями
import openpyxl # Добавили импорт openpyxl, он нам понадобится
from openpyxl import load_workbook # Для combine_excel_files
from openpyxl.utils import get_column_letter # Для combine_excel_files
from openpyxl.styles import Font, Alignment # Для combine_excel_files

# --- Импорты пользовательских обработчиков ---
# Старые пути, которые мы сейчас поменяем:
# USER_HANDLERS_BASE_PATH = Path("/Users/aliya/handlers") 
# Обновленные пути для работы в Docker (и на хосте, если папка "handlers" лежит рядом с проектом или доступна по абсолютному пути):
# Предполагаем, что папка /Users/aliya/handlers будет смонтирована в /app_external_handlers внутри контейнера
USER_HANDLERS_BASE_PATH = Path("/app_external_handlers") 
import sys

# 1. Добавляем ОСНОВНУЮ папку /app_external_handlers в sys.path
# Это позволит импортам типа `from utils import ...` внутри твоих скриптов найти /Users/aliya/handlers/utils.py
sys.path.insert(0, str(USER_HANDLERS_BASE_PATH))

# 2. Добавляем пути к конкретным папкам обработчиков в sys.path
# Это позволит Python находить сами модули обработчиков (smeta_rus_processor.py и т.д.)
sys.path.insert(0, str(USER_HANDLERS_BASE_PATH / "smeta_ru"))
sys.path.insert(0, str(USER_HANDLERS_BASE_PATH / "grandsmeta"))
sys.path.insert(0, str(USER_HANDLERS_BASE_PATH / "turbosmetchik"))

# Теперь пробуем импортировать функции
process_smeta_ru = None
try:
    from smeta_rus_processor import process_smeta_ru as psr # Ожидаем файл smeta_rus_processor.py
    process_smeta_ru = psr
    print("[TASKS_DEBUG] Успешно импортирован process_smeta_ru из smeta_rus_processor.py")
except ImportError as e:
    print(f"[TASKS_DEBUG] Ошибка импорта process_smeta_ru: {e}.")
except Exception as e:
    print(f"[TASKS_DEBUG] Неожиданная ошибка при импорте process_smeta_ru: {e}")

process_grandsmeta_mixed = None
try:
    from grandsmeta_processor import process_grandsmeta_mixed as pgm # Ожидаем файл grandsmeta_processor.py
    process_grandsmeta_mixed = pgm
    print("[TASKS_DEBUG] Успешно импортирован process_grandsmeta_mixed из grandsmeta_processor.py")
except ImportError as e:
    print(f"[TASKS_DEBUG] Ошибка импорта process_grandsmeta_mixed: {e}.")
except Exception as e:
    print(f"[TASKS_DEBUG] Неожиданная ошибка при импорте process_grandsmeta_mixed: {e}")

process_turbosmetchik_1 = None
process_turbosmetchik_2 = None # Заглушка
process_turbosmetchik_3 = None # Заглушка

try:
    from processor_1 import process_turbosmetchik_1 as pt1 # Ожидаем файл processor_1.py в turbosmetchik
    process_turbosmetchik_1 = pt1
    print("[TASKS_DEBUG] Успешно импортирован process_turbosmetchik_1")
except ImportError as e:
    print(f"[TASKS_DEBUG] Ошибка импорта process_turbosmetchik_1: {e}.")
except Exception as e:
    print(f"[TASKS_DEBUG] Неожиданная ошибка при импорте process_turbosmetchik_1: {e}")

# TODO: Добавить импорты для process_turbosmetchik_2 и process_turbosmetchik_3

# --- Конец импортов пользовательских обработчиков ---

def create_excel_from_data(output_dir_path: Path, original_file_name: str, headers: list, data_rows: list) -> str:
    """
    Создает Excel-файл из предоставленных заголовков и строк данных.

    Args:
        output_dir_path (Path): Путь к директории, куда будет сохранен Excel-файл.
        original_file_name (str): Имя оригинального файла, используется для формирования имени нового файла.
        headers (list): Список заголовков столбцов.
        data_rows (list): Список списков, где каждый внутренний список - это строка данных.

    Returns:
        str: Полный путь к созданному Excel-файлу.
    """
    # Убедимся, что директория для вывода существует
    output_dir_path.mkdir(parents=True, exist_ok=True)

    # Формируем имя выходного файла
    # Убираем расширение из original_file_name, если оно есть
    base_name, _ = os.path.splitext(original_file_name)
    output_file_name = f"processed_{base_name}.xlsx"
    output_file_path = output_dir_path / output_file_name

    # Создаем новую книгу Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active # Берем активный лист

    # Записываем заголовки
    if headers:
        sheet.append(headers)

    # Записываем строки данных
    if data_rows:
        for row in data_rows:
            sheet.append(row)

    # Сохраняем книгу
    workbook.save(output_file_path)
    print(f"[TASKS_DEBUG] Создан Excel файл: {output_file_path}")
    return str(output_file_path)

# Новая функция для объединения Excel файлов
def combine_excel_files(
    individual_excel_paths: list[str],
    output_dir_for_combined_file: Path,
    combined_file_name_base: str,
    processor_type: str
) -> str | None:
    if not individual_excel_paths:
        print("[COMBINE_DEBUG] No individual files to combine.")
        return None

    combined_wb = openpyxl.Workbook()
    combined_ws = combined_wb.active
    sheet_title = processor_type[:31] if processor_type else "Combined_Data"
    try:
        combined_ws.title = sheet_title
    except ValueError: # Если title некорректный
        combined_ws.title = "Combined_Data_Fallback"


    headers_written = False
    common_headers = []

    # Пытаемся получить общие заголовки из первого файла
    if individual_excel_paths:
        try:
            # data_only=True чтобы читать значения, а не формулы
            first_file_wb = load_workbook(individual_excel_paths[0], data_only=True)
            first_file_ws = first_file_wb.active
            if first_file_ws.max_row >= 1:
                common_headers = [cell.value for cell in first_file_ws[1]]
                if any(c is not None for c in common_headers): # Пишем заголовки, только если они не пустые
                    combined_ws.append(common_headers)
                    # Опционально: стиль для заголовков
                    for col_idx, header_value in enumerate(common_headers, 1):
                        cell = combined_ws.cell(row=1, column=col_idx)
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    headers_written = True
                else:
                    common_headers = [] # Сбрасываем, если заголовки оказались пустыми
            print(f"[COMBINE_DEBUG] Common headers: {common_headers}")
        except Exception as e:
            print(f"[COMBINE_DEBUG] Error reading headers from first file {individual_excel_paths[0]}: {e}")
            common_headers = [] # Убедимся, что это список

    is_multi_file_job = len(individual_excel_paths) > 1

    for file_idx, individual_file_path_str in enumerate(individual_excel_paths):
        individual_file_path = Path(individual_file_path_str)
        # Извлекаем "чистое" имя файла для разделителя
        original_file_basename = os.path.splitext(individual_file_path.name.replace("processed_", ""))[0]

        if is_multi_file_job:
            # Добавляем пустую строку перед разделителем, если в листе уже что-то есть
            # (например, общие заголовки или данные предыдущего файла)
            if combined_ws.max_row > (1 if headers_written and file_idx == 0 else 0) :
                 combined_ws.append([])

            separator_text = f"Данные из файла: {original_file_basename}"
            sep_row_idx = combined_ws.max_row + 1 # Следующая доступная строка
            combined_ws.cell(row=sep_row_idx, column=1, value=separator_text)

            num_cols_to_merge_sep = len(common_headers) if common_headers else 5 # Объединяем на ширину заголовков или 5 колонок
            if num_cols_to_merge_sep <= 0: num_cols_to_merge_sep = 1 # Минимум 1 колонка

            try:
                combined_ws.merge_cells(
                    start_row=sep_row_idx, start_column=1,
                    end_row=sep_row_idx, end_column=num_cols_to_merge_sep
                )
            except Exception as merge_err:
                 print(f"[COMBINE_DEBUG] Error merging separator cells: {merge_err}. Cols to merge: {num_cols_to_merge_sep}")

            cell_to_format = combined_ws.cell(row=sep_row_idx, column=1)
            cell_to_format.font = Font(bold=True, color="003366") # Темно-синий цвет для акцента
            cell_to_format.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            # Пустая строка после разделителя для лучшего визуального разделения
            # combined_ws.append([])


        try:
            source_wb = load_workbook(individual_file_path_str, data_only=True)
            source_ws = source_wb.active

            # Если общие заголовки были записаны, начинаем читать со второй строки исходных файлов.
            # Иначе - с первой.
            start_row_source = 2 if headers_written else 1

            for row_num_source in range(start_row_source, source_ws.max_row + 1):
                row_data = [source_ws.cell(row=row_num_source, column=col_num_source).value
                            for col_num_source in range(1, source_ws.max_column + 1)]

                if not any(c is not None for c in row_data): # Пропускаем полностью пустые строки
                    continue

                # Если есть общие заголовки, подгоняем количество столбцов в данных
                if common_headers:
                    if len(row_data) < len(common_headers):
                        row_data.extend([None] * (len(common_headers) - len(row_data)))
                    elif len(row_data) > len(common_headers):
                        row_data = row_data[:len(common_headers)]
                
                # Записываем данные и применяем wrap_text к каждой ячейке
                current_row_in_combined = combined_ws.max_row + 1
                for col_idx_data, cell_value in enumerate(row_data, 1):
                    cell = combined_ws.cell(row=current_row_in_combined, column=col_idx_data, value=cell_value)
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                
                written_row_idx_in_combined = current_row_in_combined # Обновляем, так как теперь сами управляем записью строки

                is_footer_row = row_data and row_data[0] == "__FOOTER__"

                if is_footer_row:
                    footer_content = row_data[1] if len(row_data) > 1 and row_data[1] is not None else "Итог"
                    # Перезаписываем первую ячейку футера (она уже была создана циклом выше)
                    cell_footer_text = combined_ws.cell(row=written_row_idx_in_combined, column=1, value=footer_content)
                    # Применяем стиль к ячейке с текстом футера
                    cell_footer_text.font = Font(bold=True)
                    cell_footer_text.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)

                    max_cols_footer = len(common_headers) if common_headers else (len(row_data) if len(row_data) > 1 else 1)
                    if max_cols_footer <= 0: max_cols_footer = 1
                    
                    # Очищаем остальные ячейки в строке футера
                    for col_idx_clear in range(2, max_cols_footer + 1):
                        combined_ws.cell(row=written_row_idx_in_combined, column=col_idx_clear, value=None)
                    
                    try:
                        combined_ws.merge_cells(
                            start_row=written_row_idx_in_combined, start_column=1,
                            end_row=written_row_idx_in_combined, end_column=max_cols_footer
                        )
                    except Exception as merge_err:
                        print(f"[COMBINE_DEBUG] Error merging footer cells: {merge_err}. Cols: {max_cols_footer}")
                    
                    cell_to_style = combined_ws.cell(row=written_row_idx_in_combined, column=1)
                    cell_to_style.font = Font(bold=True)
                    cell_to_style.alignment = Alignment(horizontal='right') # Итоги обычно справа
                else: # Обычная строка данных
                    if len(row_data) >= 3 and row_data[1] is not None and row_data[2] is None:
                        # Объединяем 2-ю и 3-ю колонки, если есть хотя бы 3 колонки по заголовкам/данным
                        min_cols_for_merge_check = len(common_headers) if common_headers else len(row_data)
                        if min_cols_for_merge_check >=3:
                            try:
                                combined_ws.merge_cells(
                                    start_row=written_row_idx_in_combined, start_column=2,
                                    end_row=written_row_idx_in_combined, end_column=3
                                )
                                # Для объединенной ячейки тоже можно установить alignment, если нужно
                                merged_cell_data = combined_ws.cell(row=written_row_idx_in_combined, column=2)
                                merged_cell_data.alignment = Alignment(wrap_text=True, vertical='top')

                            except Exception as merge_err:
                                print(f"[COMBINE_DEBUG] Error merging data cells [2,3]: {merge_err}")
        except Exception as e_file:
            print(f"[COMBINE_DEBUG] Error processing data from {individual_file_path_str}: {e_file}")
            combined_ws.append([f"Ошибка чтения данных из файла: {individual_file_path.name}", str(e_file)])

    # --- НАЧАЛО ЛОГИКИ УСТАНОВКИ ШИРИНЫ КОЛОНОК ---
    reference_widths_list = None
    # Старый путь:
    # REF_FILES_BASE_PATH = Path("/Users/aliya/handlers/reference_files")
    # Новый путь для Docker:
    REF_FILES_BASE_PATH = USER_HANDLERS_BASE_PATH / "reference_files" # Используем уже определенный USER_HANDLERS_BASE_PATH

    REF_SMETA_RU_PATH = REF_FILES_BASE_PATH / "Смета ру.xlsm"
    REF_TURBO_PATH = REF_FILES_BASE_PATH / "Турбосметчик1,2,3.xlsm"
    REF_GRAND_PATH = REF_FILES_BASE_PATH / "Пример1 2.xlsx"

    selected_ref_path = None
    if processor_type == 'smeta_ru':
        selected_ref_path = REF_SMETA_RU_PATH
    elif processor_type == 'turbosmetchik': # Для всех подтипов Турбосметчика
        selected_ref_path = REF_TURBO_PATH
    elif processor_type == 'grand_smeta':
        selected_ref_path = REF_GRAND_PATH

    if selected_ref_path and selected_ref_path.exists():
        try:
            wb_ref = load_workbook(selected_ref_path, data_only=True)
            ws_ref = wb_ref.active
            # Читаем ширины для первых 6 колонок (A-F)
            # Используем default=8.43, если .width не установлен (None)
            reference_widths_list = []
            for i in range(1, 7): # Колонки с 1 по 6 (A-F)
                col_dim = ws_ref.column_dimensions[get_column_letter(i)]
                # Проверяем, что col_dim.width это число, иначе 8.43
                width_val = col_dim.width if isinstance(col_dim.width, (int, float)) and col_dim.width > 0 else 8.43
                reference_widths_list.append(width_val)
            wb_ref.close()
            print(f"[COMBINE_DEBUG] Успешно прочитаны эталонные ширины из {selected_ref_path.name}: {reference_widths_list}")
        except Exception as e_ref:
            print(f"[COMBINE_DEBUG] Не удалось прочитать эталонные ширины из {selected_ref_path.name}: {e_ref}")
            reference_widths_list = None # Сбрасываем, если ошибка
    else:
        if selected_ref_path:
            print(f"[COMBINE_DEBUG] Эталонный файл не найден: {selected_ref_path}")
        else:
            print(f"[COMBINE_DEBUG] Эталонный файл не определен для типа обработчика: {processor_type}")

    # Применяем ширины
    if reference_widths_list:
        print(f"[COMBINE_DEBUG] Применение эталонных ширин: {reference_widths_list}")
        for i, width_val in enumerate(reference_widths_list):
            col_letter = get_column_letter(i + 1) # i+1 т.к. список с 0, а колонки с 1
            combined_ws.column_dimensions[col_letter].width = width_val
        
        # Для колонок ПОСЛЕ эталонных (если они есть) - автоподбор
        num_ref_cols = len(reference_widths_list) # Должно быть 6
        max_cols_in_data = combined_ws.max_column

        if max_cols_in_data > num_ref_cols:
            print(f"[COMBINE_DEBUG] Автоподбор для колонок с {get_column_letter(num_ref_cols + 1)} до {get_column_letter(max_cols_in_data)}")
            for col_idx in range(num_ref_cols + 1, max_cols_in_data + 1):
                col_letter_auto = get_column_letter(col_idx)
                max_length_auto = 0
                for cell_auto in combined_ws[col_letter_auto]: # Проход по всем ячейкам в данной колонке
                    try:
                        if cell_auto.value:
                            cell_text_length_auto = len(str(cell_auto.value))
                            if cell_text_length_auto > max_length_auto:
                                max_length_auto = cell_text_length_auto
                    except:
                        pass # Пропускаем ошибки получения длины
                adjusted_width_auto = (max_length_auto + 2) if max_length_auto > 0 else 8.43
                if adjusted_width_auto > 60: adjusted_width_auto = 60 # Ограничение максимальной ширины
                combined_ws.column_dimensions[col_letter_auto].width = adjusted_width_auto
    else:
        print("[COMBINE_DEBUG] Эталонные ширины не загружены. Применяется автоподбор ко всем колонкам.")
        # Если эталонные ширины не загружены, применяем автоподбор ко всем колонкам
        for col in combined_ws.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column) # col[0] - первая ячейка в колонке
            for cell in col:
                try:
                    if cell.value:
                        cell_text_length = len(str(cell.value))
                        if cell_text_length > max_length:
                            max_length = cell_text_length
                except:
                    pass # Пропускаем, если не можем получить длину
            adjusted_width = (max_length + 2) if max_length > 0 else 8.43 # мин. ширина 8.43
            if adjusted_width > 60: adjusted_width = 60 # макс. ширина 60
            combined_ws.column_dimensions[column_letter].width = adjusted_width
    # --- КОНЕЦ ЛОГИКИ УСТАНОВКИ ШИРИНЫ КОЛОНОК ---

    output_dir_for_combined_file.mkdir(parents=True, exist_ok=True)
    # Убедимся, что имя файла имеет расширение .xlsx
    base, ext = os.path.splitext(combined_file_name_base)
    final_combined_file_name = f"{base}_processed.xlsx"

    full_combined_path = output_dir_for_combined_file / final_combined_file_name

    try:
        combined_wb.save(full_combined_path)
        print(f"[COMBINE_DEBUG] Combined Excel file saved to: {full_combined_path}")
        return str(full_combined_path)
    except Exception as e_save:
        print(f"[COMBINE_DEBUG] Error saving combined Excel file '{full_combined_path}': {e_save}")
        return None


@shared_task(bind=True)
def process_uploaded_file(self, file_path, processor_type, sub_type=None, db_task_id=None):
    """
    Основная задача Celery для обработки загруженного файла.
    Распаковывает zip, если это zip, считает файлы, обрабатывает каждый, обновляя статус.
    """
    
    # 1. Получаем объект ProcessingTask из БД
    try:
        task_db_instance = ProcessingTask.objects.get(id=db_task_id)
    except ProcessingTask.DoesNotExist:
        print(f"КРИТИЧЕСКАЯ ОШИБКА: Задача с ID {db_task_id} не найдена в БД в Celery worker.")
        # В этом случае worker не может обновить статус, так как нет объекта.
        # Можно попробовать создать self.update_state с ошибкой, но это не повлияет на БД.
        # Лучше всего, если такая ситуация вообще не будет возникать.
        return

    task_db_instance.status = 'STARTED' 
    task_db_instance.save(update_fields=['status', 'updated_at'])

    # 2. Обновляем начальное состояние прогресса в Celery и в БД
    initial_status_message = f'Начало анализа файла {os.path.basename(file_path)}...'
    self.update_state(state='PROGRESS', meta={
        'current': 0, 
        'total': 1, # По умолчанию, пока не знаем точное количество
        'status_message': initial_status_message
    })
    task_db_instance.update_progress(current=0, total=1, status_message=initial_status_message)

    files_to_process_paths = [] # Список путей (Path objects) к файлам, которые нужно обработать
    temp_extract_dir = None    # Путь к временной директории, если это zip
    processed_individual_excel_paths = [] # Список путей к индивидуально обработанным Excel файлам

    # Директория для сохранения результатов этой конкретной задачи
    # processed_tasks -> task_id -> individual_results -> processed_file1.xlsx, processed_file2.xlsx
    #                       -> final_combined_result.xlsx (позже)
    output_dir_for_individual_files = Path(settings.MEDIA_ROOT) / 'processed_tasks' / str(db_task_id) / 'individual_results'
    try:
        output_dir_for_individual_files.mkdir(parents=True, exist_ok=True)
    except Exception as e:
        error_message = f"Не удалось создать директорию для результатов: {output_dir_for_individual_files}. Ошибка: {str(e)}"
        self.update_state(state='FAILURE', meta={'current': 0, 'total': 0, 'status_message': error_message})
        task_db_instance.mark_as_failed(error_message)
        return

    # 3. Проверяем, zip ли это, и подготавливаем список файлов для обработки
    original_file_path = Path(file_path) # Преобразуем строку пути в объект Path

    if original_file_path.suffix.lower() == '.zip':
        try:
            temp_extract_base = Path(settings.MEDIA_ROOT) / 'temp_unzip_tasks'
            temp_extract_dir = temp_extract_base / str(db_task_id)
            temp_extract_dir.mkdir(parents=True, exist_ok=True)

            with zipfile.ZipFile(original_file_path, 'r') as zip_ref:
                found_any_processable_file = False
                for info in zip_ref.infolist():
                    if info.is_dir(): # Пропускаем директории внутри архива
                        continue

                    original_filename_in_zip = info.filename # Имя файла, как его вернул zipfile
                    corrected_filename_str = original_filename_in_zip # По умолчанию, если коррекция не удастся

                    try:
                        # Применяем предложенный способ коррекции:
                        # Если имя в UTF-8, но было ошибочно прочитано как CP437,
                        # эта операция вернет его к правильному UTF-8 виду.
                        temp_bytes = original_filename_in_zip.encode('cp437')
                        corrected_filename_str = temp_bytes.decode('utf-8', 'ignore')
                        if corrected_filename_str != original_filename_in_zip:
                            print(f"[TASKS_DEBUG] Применена коррекция имени файла: '{original_filename_in_zip}' -> '{corrected_filename_str}'")
                    except UnicodeEncodeError:
                        # Это может случиться, если original_filename_in_zip содержит символы, которых нет в cp437.
                        # Значит, наша теория о "UTF-8 прочитано как CP437" неверна для этого файла.
                        print(f"[TASKS_DEBUG] Имя файла '{original_filename_in_zip}' вызвало UnicodeEncodeError при кодировании в cp437. Используется оригинальное имя.")
                    except Exception as e_filename_corr:
                        print(f"[TASKS_DEBUG] Ошибка при коррекции имени файла '{original_filename_in_zip}': {e_filename_corr}. Используется оригинальное имя.")

                    # Извлекаем файл (он извлечется с original_filename_in_zip)
                    try:
                        zip_ref.extract(info, temp_extract_dir)
                    except Exception as e_extract:
                        print(f"[TASKS_DEBUG] Ошибка извлечения файла '{info.filename}' из ZIP: {e_extract}. Пропускаем.")
                        continue
                    
                    path_on_disk_after_extraction = temp_extract_dir / original_filename_in_zip
                    final_path_for_processing = path_on_disk_after_extraction # По умолчанию

                    if corrected_filename_str != original_filename_in_zip:
                        # Если имя было скорректировано, пытаемся переименовать файл на диске
                        corrected_path_on_disk = temp_extract_dir / corrected_filename_str
                        corrected_path_on_disk.parent.mkdir(parents=True, exist_ok=True) # Убедимся, что родительская папка есть
                        try:
                            if path_on_disk_after_extraction.exists():
                                os.rename(path_on_disk_after_extraction, corrected_path_on_disk)
                                print(f"[TASKS_DEBUG] Файл переименован в: {corrected_path_on_disk}")
                                final_path_for_processing = corrected_path_on_disk
                            elif corrected_path_on_disk.exists(): # Если уже есть файл с таким именем (маловероятно)
                                print(f"[TASKS_DEBUG] Файл уже существует по пути: {corrected_path_on_disk}")
                                final_path_for_processing = corrected_path_on_disk
                            else:
                                print(f"[TASKS_DEBUG] Исходный извлеченный файл {path_on_disk_after_extraction} не найден для переименования. Используем оригинальное имя, если файл существует.")
                                # final_path_for_processing остается path_on_disk_after_extraction
                        except Exception as e_rename:
                            print(f"[TASKS_DEBUG] Не удалось переименовать '{path_on_disk_after_extraction}' в '{corrected_path_on_disk}': {e_rename}. Используем оригинальное извлеченное имя, если файл существует.")
                            # final_path_for_processing остается path_on_disk_after_extraction
                    
                    # Проверяем, существует ли файл по итоговому пути
                    if not final_path_for_processing.exists() or not final_path_for_processing.is_file():
                        print(f"[TASKS_DEBUG] Файл {final_path_for_processing}, предназначенный для обработки, не существует или не является файлом. Пропускаем.")
                        continue

                    # Фильтрация служебных файлов и файлов неверного типа (уже по corrected_filename_str или final_path_for_processing.name)
                    filename_for_check = final_path_for_processing.name
                    full_path_str_for_check = str(final_path_for_processing)

                    # Сначала проверяем на временные файлы Office, так как они могут иметь расширение .xlsx/.xls
                    if filename_for_check.startswith('~$'):
                        print(f"[TASKS_DEBUG] Пропускаем временный файл MS Office: {filename_for_check}")
                        try: final_path_for_processing.unlink(missing_ok=True) # Удаляем его из временной папки
                        except Exception as e_del_temp: print(f"[TASKS_DEBUG] Ошибка при удалении временного файла {final_path_for_processing}: {e_del_temp}")
                        continue # Переходим к следующему файлу в архиве

                    if "__MACOSX/" in full_path_str_for_check or filename_for_check.startswith("._"):
                        print(f"[TASKS_DEBUG] Пропускаем служебный файл macOS: {filename_for_check}")
                        try: final_path_for_processing.unlink(missing_ok=True)
                        except Exception as e_del_junk: print(f"[TASKS_DEBUG] Ошибка при удалении служебного файла {final_path_for_processing}: {e_del_junk}")
                        continue

                    if not filename_for_check.lower().endswith(('.xlsx', '.xls')):
                        print(f"[TASKS_DEBUG] Пропускаем файл с неподдерживаемым расширением: {filename_for_check}")
                        try: final_path_for_processing.unlink(missing_ok=True)
                        except Exception as e_del_junk_ext: print(f"[TASKS_DEBUG] Ошибка при удалении файла с неподдерживаемым расширением {final_path_for_processing}: {e_del_junk_ext}")
                        continue
                    
                    files_to_process_paths.append(final_path_for_processing)
                    found_any_processable_file = True
                
                if not found_any_processable_file:
                    error_message = 'Zip-архив не содержит подходящих файлов Excel для обработки или все файлы являются служебными.'
                    self.update_state(state='FAILURE', meta={'current': 0, 'total': 0, 'status_message': error_message})
                    task_db_instance.mark_as_failed(error_message)
                    if temp_extract_dir.exists(): # Очистка, если папка была создана
                        import shutil
                        try: shutil.rmtree(temp_extract_dir)
                        except Exception as e_shutil_empty: print(f"[TASKS_DEBUG] Ошибка при удалении {temp_extract_dir} (нет файлов): {e_shutil_empty}")
                    return
            
            total_files = len(files_to_process_paths)
            status_message_after_unzip = f'Распаковано и подготовлено {total_files} файлов из {original_file_path.name}. Начинаю обработку...'
            self.update_state(state='PROGRESS', meta={'current': 0, 'total': total_files, 'status_message': status_message_after_unzip})
            task_db_instance.update_progress(0, total_files, status_message_after_unzip)

        except zipfile.BadZipFile:
            error_message = 'Ошибка: Неверный формат zip-архива.'
            self.update_state(state='FAILURE', meta={'current': 0, 'total': 0, 'status_message': error_message})
            task_db_instance.mark_as_failed(error_message)
            return
        except Exception as e:
            error_message = f'Ошибка при распаковке zip-архива: {str(e)}'
            self.update_state(state='FAILURE', meta={'current': 0, 'total': 0, 'status_message': error_message})
            task_db_instance.mark_as_failed(error_message)
            return
    else:
        # Если это не zip, а одиночный файл
        files_to_process_paths.append(original_file_path)
        total_files = 1
        status_message_single_file = f'Подготовка к обработке одиночного файла: {original_file_path.name}'
        self.update_state(state='PROGRESS', meta={'current': 0, 'total': total_files, 'status_message': status_message_single_file})
        task_db_instance.update_progress(0, total_files, status_message_single_file)

    # 4. Основной цикл обработки файлов
    if not files_to_process_paths: # Дополнительная проверка, если вдруг список пуст
        error_message = "Не найдено файлов для обработки после анализа."
        self.update_state(state='FAILURE', meta={'current': 0, 'total': 0, 'status_message': error_message})
        task_db_instance.mark_as_failed(error_message)
        return
        
    total_files = len(files_to_process_paths) # Обновляем total_files на случай если он изменился

    for i, current_file_to_process_path in enumerate(files_to_process_paths):
        current_file_number = i + 1
        file_name_for_status_update = current_file_to_process_path.name 
        
        status_message_processing_file = f'Обработка файла {current_file_number} из {total_files}: {file_name_for_status_update} (Тип: {processor_type})'
        self.update_state(state='PROGRESS', meta={
            'current': current_file_number, 
            'total': total_files, 
            'status_message': status_message_processing_file
        })
        task_db_instance.update_progress(current_file_number, total_files, status_message_processing_file)
        
        try:
            headers = None
            data_rows = None
            
            # Выбираем и вызываем соответствующий пользовательский обработчик
            if processor_type == 'smeta_ru':
                if process_smeta_ru:
                    print(f"[TASKS_DEBUG] Вызов process_smeta_ru для {str(current_file_to_process_path)}")
                    headers, data_rows = process_smeta_ru(str(current_file_to_process_path))
                else:
                    raise ValueError("Обработчик process_smeta_ru не импортирован.")
            elif processor_type == 'grand_smeta':
                if process_grandsmeta_mixed:
                    print(f"[TASKS_DEBUG] Вызов process_grandsmeta_mixed для {str(current_file_to_process_path)}")
                    headers, data_rows = process_grandsmeta_mixed(str(current_file_to_process_path))
                else:
                    raise ValueError("Обработчик process_grandsmeta_mixed не импортирован.")
            elif processor_type == 'turbosmetchik':
                # Здесь может быть логика выбора подтипа, если process_turbosmetchik_1, _2, _3 разные
                if sub_type == 'type1' or not sub_type: # По умолчанию или если явно type1
                    if process_turbosmetchik_1:
                        print(f"[TASKS_DEBUG] Вызов process_turbosmetchik_1 для {str(current_file_to_process_path)}")
                        headers, data_rows = process_turbosmetchik_1(str(current_file_to_process_path))
                    else:
                        raise ValueError("Обработчик process_turbosmetchik_1 не импортирован.")
                # Добавить elif для sub_type == 'type2' и process_turbosmetchik_2 и т.д.
                else:
                    raise ValueError(f"Неизвестный или неимпортированный подтип обработчика для Турбосметчик: {sub_type}")
            else:
                raise ValueError(f"Неизвестный тип основного обработчика: {processor_type}")

            # Проверяем, что данные получены
            if headers is None or data_rows is None:
                # Это может случиться, если скрипт отработал, но вернул не то, что ожидалось
                # или не вернул ничего (например, пустой файл на входе).
                # Решим, считать ли это ошибкой или просто пропустить файл.
                # Пока считаем, что если нет данных, то это проблема.
                raise ValueError(f"Обработчик для {file_name_for_status_update} не вернул данные (headers или data_rows).")

            # Создаем Excel файл из полученных данных
            print(f"[TASKS_DEBUG] Вызов create_excel_from_data для {file_name_for_status_update}")
            processed_excel_path = create_excel_from_data(
                output_dir_path=output_dir_for_individual_files,
                original_file_name=file_name_for_status_update, # Используем имя текущего обрабатываемого файла
                headers=headers,
                data_rows=data_rows
            )
            processed_individual_excel_paths.append(processed_excel_path)
            print(f"[TASKS_DEBUG] Успешно обработан и сохранен: {processed_excel_path}")

        except Exception as e_file_processing:
            error_message_file = f"Критическая ошибка при обработке файла {file_name_for_status_update}: {str(e_file_processing)}"
            # Обновляем статус всей задачи на FAILURE
            self.update_state(state='FAILURE', meta={'current': current_file_number, 'total': total_files, 'status_message': error_message_file})
            task_db_instance.mark_as_failed(error_message_file)
            # Очищаем временную директорию, если она была создана для ZIP
            if temp_extract_dir and temp_extract_dir.exists():
                import shutil
                try:
                    shutil.rmtree(temp_extract_dir)
                    print(f"[TASKS_DEBUG] Временная папка {temp_extract_dir} удалена после ошибки.")
                except Exception as e_shutil:
                    print(f"[TASKS_DEBUG] Ошибка при удалении временной папки {temp_extract_dir} после ошибки: {e_shutil}")
            return # Завершаем всю задачу, так как один из файлов не обработался

    # 5. Завершение задачи и ОБЪЕДИНЕНИЕ ФАЙЛОВ
    if not processed_individual_excel_paths:
        # Сюда мы можем попасть, если files_to_process_paths был пуст изначально (хотя есть проверка выше)
        # или если все файлы в цикле вызвали ошибку, которая не привела к return (маловероятно с текущей логикой)
        # или если не было файлов для обработки вообще.
        final_message = "Не было создано ни одного индивидуально обработанного файла. Объединение не требуется."
        self.update_state(state='FAILURE', meta={'current': 0, 'total': total_files, 'status_message': final_message}) # total_files может быть 0
        task_db_instance.mark_as_failed(final_message)
        # Очищаем временную директорию от ZIP, если она была
        if temp_extract_dir and temp_extract_dir.exists():
            import shutil
            try:
                shutil.rmtree(temp_extract_dir)
            except Exception as e_shutil_fail:
                 print(f"[TASKS_DEBUG] Ошибка при удалении временной папки {temp_extract_dir} при общем сбое: {e_shutil_fail}")
        return

    # ПЫТАЕМСЯ ОБЪЕДИНИТЬ ФАЙЛЫ
    combined_excel_path = None
    # Директория для сохранения итогового объединенного файла (на уровень выше individual_results)
    final_output_dir = Path(settings.MEDIA_ROOT) / 'processed_tasks' / str(db_task_id)
    
    # Получаем базовое имя оригинального загруженного файла для имени объединенного файла
    base_original_uploaded_name, _ = os.path.splitext(task_db_instance.original_file_name)

    try:
        print(f"[TASKS_DEBUG] Попытка объединить {len(processed_individual_excel_paths)} файлов в один.")
        combined_excel_path = combine_excel_files(
            individual_excel_paths=processed_individual_excel_paths,
            output_dir_for_combined_file=final_output_dir,
            combined_file_name_base=base_original_uploaded_name,
            processor_type=processor_type # Передаем тип обработчика для именования листа
        )
    except Exception as e_combine_call:
        print(f"[TASKS_DEBUG] Критическая ошибка при вызове combine_excel_files: {e_combine_call}")
        # Не будем считать это провалом всей задачи, если индивидуальные файлы есть,
        # но запишем в статус, что объединение не удалось.

    path_for_result_link = None # Абсолютный путь к файлу, который будет результатом задачи
    relative_result_path_for_frontend = None # Путь относительно MEDIA_ROOT для ссылки на фронте

    if combined_excel_path:
        final_success_message = (
            f'Все {len(processed_individual_excel_paths)} из {total_files} файлов в задаче '
            f'"{task_db_instance.original_file_name}" были обработаны. '
            f'Результаты объединены в файл: {os.path.basename(combined_excel_path)}.'
        )
        path_for_result_link = combined_excel_path
        result_data_details = {
            'combined_file_path': combined_excel_path, # Абсолютный путь
            'individual_files_location': str(output_dir_for_individual_files),
            'total_files_processed': len(processed_individual_excel_paths),
            'individual_files_paths': processed_individual_excel_paths
        }
    else:
        final_success_message = (
            f'{len(processed_individual_excel_paths)} из {total_files} файлов в задаче '
            f'"{task_db_instance.original_file_name}" были индивидуально обработаны. '
            f'Объединить файлы в один не удалось или не требовалось (например, был обработан только один файл, и он же является результатом, или произошла ошибка при объединении). '
            f'Индивидуальные файлы (если есть) находятся в папке: {str(output_dir_for_individual_files.name)}.'
        )
        # Если объединение не удалось, но есть индивидуальные файлы, ссылаемся на первый из них
        if processed_individual_excel_paths:
            path_for_result_link = processed_individual_excel_paths[0]
        
        result_data_details = {
            'error_or_skipped_combining_files': True,
            'info_message': 'Объединенный файл не создан. Результатом является первый индивидуально обработанный файл (если есть) или смотрите папку с индивидуальными результатами.',
            'individual_files_location': str(output_dir_for_individual_files),
            'total_files_processed': len(processed_individual_excel_paths),
            'individual_files_paths': processed_individual_excel_paths
        }
    
    # Получаем относительный путь для result_link_path в meta, если path_for_result_link существует
    if path_for_result_link:
        try:
            abs_path_obj = Path(path_for_result_link)
            media_root_path = Path(settings.MEDIA_ROOT)
            # Преобразуем в строку и заменяем обратные слеши на прямые для URL-совместимости
            relative_result_path_for_frontend = str(abs_path_obj.relative_to(media_root_path)).replace("\\\\", "/")
        except ValueError:
            print(f"[TASKS_DEBUG] Не удалось получить относительный путь для {path_for_result_link} относительно {settings.MEDIA_ROOT}")
            relative_result_path_for_frontend = None # Оставляем None, если не получается

    self.update_state(state='SUCCESS', meta={
        'current': len(processed_individual_excel_paths),
        'total': total_files,
        'status_message': final_success_message,
        'result_link_path': relative_result_path_for_frontend, # Относительный путь для фронтенда
        'result_absolute_path': path_for_result_link, # Абсолютный путь для модели/бекенда
        'result_details': result_data_details
    })
    
    # В модель ProcessingTask мы сохраняем абсолютный путь к основному файлу результата
    task_db_instance.mark_as_success(
        final_success_message,
        result_file_path_param=path_for_result_link, # Абсолютный путь
        result_data=result_data_details
    )

    # Очистка временной папки (если это был ZIP) ПОСЛЕ успешной обработки всех его частей
    if temp_extract_dir and temp_extract_dir.exists():
        import shutil
        try:
            shutil.rmtree(temp_extract_dir)
            print(f"[TASKS_DEBUG] Временная папка {temp_extract_dir} успешно удалена после завершения задачи.")
        except Exception as e_shutil_final:
            print(f"[TASKS_DEBUG] Ошибка при удалении временной папки {temp_extract_dir} в конце: {e_shutil_final}")

    return {'final_status': final_success_message, 'total_processed': len(processed_individual_excel_paths), 'final_output_path': path_for_result_link} 