import openpyxl
import traceback
# Используем АБСОЛЮТНЫЙ импорт utils
from utils import is_likely_empty, check_merge, get_start_coord, get_item_id_nature

# Вспомогательная функция для проверки на 0 или пустоту
def is_zero_or_empty(value):
    if is_likely_empty(value):
        return True
    try:
        # Попытка преобразовать в float и сравнить с 0
        # Убираем пробелы и заменяем запятую для чисел типа ' 123,45 '
        numeric_value = float(str(value).replace(',', '.').strip())
        return numeric_value == 0
    except (ValueError, TypeError):
        # Если не удалось преобразовать в число, считаем не нулем (т.к. не пустое)
        return False

# -----------------------------------------------------------
#  ➜  ПРОВЕРКА «1 2 3 … 11» — строка‑нумератор таблицы
# -----------------------------------------------------------
def is_column_number_row(row_cells, max_check=11):
    """
    Возвращает True, если первые `max_check` ячеек содержат подряд
    числа 1..max_check, а остальные ячейки пусты.
    """
    for idx in range(max_check):
        if idx >= len(row_cells):
            return False
        # Сравниваем как строки, чтобы избежать проблем с типами
        cell_val = str(row_cells[idx].value).strip() if row_cells[idx] else ""
        if cell_val != str(idx + 1):
            return False
    for cell in row_cells[max_check:]:
        if not is_likely_empty(cell.value):
            return False
    return True

def process_grandsmeta_mixed(input_path): # Изменил имя функции для новой версии
    """
    ОБРАБАТЫВАЕТ один Excel файл по НОВЫМ ПРАВИЛАМ "Турбосметчик-1" (версия 3).
    - Название раздела/подраздела идет отдельной строкой.
    - Итог по разделу/подразделу идет отдельной строкой после всех позиций.
    ВОЗВРАЩАЕТ данные (заголовки и координаты) для дальнейшей обработки.
    """
    output_headers = ["№№ п/п", "Шифр расценки и коды ресурсов", "Наименование работ и затрат", "Единица измерения", "Кол-во единиц", "ВСЕГО затрат, руб."]
    start_id_col_idx = 0    # A
    item_total_cost_col_idx = 10   # K - Итоговая цена теперь в колонке K (0-based index 10)

    processed_rows_list = []
    active_items_buffer = [] # Буфер для основных позиций, ожидающих общую цену
    # pending_section_header и pending_subsection_header больше не нужны в прежнем виде
    first_section_found = False

    workbook = None
    try:
        workbook = openpyxl.load_workbook(filename=input_path, data_only=True)
        if not workbook.sheetnames:
            return None, None
        worksheet = workbook[workbook.sheetnames[0]]

        for row_num, row_cells_tuple in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
            row_cells = list(row_cells_tuple)

            # ─── 1. ПРОПУСКАЕМ строку «1 2 3 … 11» ──────────────────────
            if is_column_number_row(row_cells):
                continue

            non_empty_info = [(i, getattr(c, 'coordinate', None), c.value) for i, c in enumerate(row_cells) if not is_likely_empty(c.value)]
            if not non_empty_info:
                continue

            cell_A = row_cells[start_id_col_idx] if len(row_cells) > start_id_col_idx else None
            cell_A_value_str = str(getattr(cell_A, 'value', '')).strip() if cell_A else ""
            cell_C = row_cells[2] if len(row_cells) > 2 else None
            cell_C_value_str = str(getattr(cell_C, 'value', '')).strip() if cell_C else ""

            row_type = None
            # Проверяем объединения для определения типа
            header_merge_AK_coord = check_merge(worksheet, row_num, 0, 10) # A(0) - K(10)
            footer_merge_CH_coord_candidate = check_merge(worksheet, row_num, 2, 7) # C(2) - H(7)

            if header_merge_AK_coord:
                if cell_A_value_str.startswith("Раздел"):
                    row_type = "section_header_name" # Новый тип
                else:
                    row_type = "subsection_header_name" # Новый тип
            elif footer_merge_CH_coord_candidate:
                if cell_C_value_str.startswith("Итого по разделу"):
                    row_type = "section_footer_total" # Новый тип
                elif cell_C_value_str.startswith("Итого по подразделу"):
                    row_type = "subsection_footer_total" # Новый тип
            elif cell_C_value_str == "Всего по позиции":
                row_type = "item_price_row"
            else:
                if cell_A and cell_A.data_type != 'f' and not is_likely_empty(cell_A.value):
                    try:
                        float(str(cell_A.value).replace(',', '.').strip())
                        row_type = "item"
                    except (ValueError, TypeError):
                        pass
            
            # Сброс буфера позиций при встрече любого хедера или футера
            if row_type in ["section_header_name", "subsection_header_name", 
                            "section_footer_total", "subsection_footer_total"] and active_items_buffer:
                if first_section_found:
                    processed_rows_list.extend(active_items_buffer)
                active_items_buffer = []

            # ─── 2. ДОПУСКАЕМ ФАЙЛЫ БЕЗ ЗАГОЛОВКОВ ─────────────────────
            if not first_section_found:
                if row_type == "item":
                    # Заголовков ещё не было, но встретилась позиция –
                    # считаем, что смета началась
                    first_section_found = True
                elif row_type != "section_header_name":
                    # Пока не встретили ни раздел, ни позицию – пропуск
                    continue
            
            # --- Обработка типов строк ---
            if row_type == "section_header_name":
                first_section_found = True # Активируем флаг
                processed_rows_list.append({
                    "type": "header_name",
                    "level": "section",
                    "source_row_num": row_num,
                    "A_K_merge_coord": header_merge_AK_coord,
                    "name_text": cell_A_value_str
                })
            elif row_type == "subsection_header_name":
                processed_rows_list.append({
                    "type": "header_name",
                    "level": "subsection",
                    "source_row_num": row_num,
                    "A_K_merge_coord": header_merge_AK_coord,
                    "name_text": cell_A_value_str
                })
            elif row_type == "section_footer_total":
                footer_price_col_idx = 10 # Колонка K (была I - 8)
                cell_K_footer = row_cells[footer_price_col_idx] if len(row_cells) > footer_price_col_idx else None # Была cell_I_footer
                processed_rows_list.append({
                    "type": "header_footer",
                    "level": "section",
                    "source_row_num": row_num,
                    "C_H_merge_coord": footer_merge_CH_coord_candidate,
                    "footer_text_content": cell_C_value_str,
                    "total_K_coord": getattr(cell_K_footer, 'coordinate', None) # Была total_I_coord
                })
            elif row_type == "subsection_footer_total":
                footer_price_col_idx = 10 # Колонка K (была I - 8)
                cell_K_footer = row_cells[footer_price_col_idx] if len(row_cells) > footer_price_col_idx else None # Была cell_I_footer
                processed_rows_list.append({
                    "type": "header_footer",
                    "level": "subsection",
                    "source_row_num": row_num,
                    "C_H_merge_coord": footer_merge_CH_coord_candidate,
                    "footer_text_content": cell_C_value_str,
                    "total_K_coord": getattr(cell_K_footer, 'coordinate', None) # Была total_I_coord
                })
            elif row_type == "item_price_row":
                cell_K_price = row_cells[item_total_cost_col_idx] if len(row_cells) > item_total_cost_col_idx else None # Changed V to K
                # Проверяем, есть ли значение в ячейке цены и не равно ли оно 0
                if cell_K_price and not is_zero_or_empty(cell_K_price.value):
                    price_total_coord_for_buffer = getattr(cell_K_price, 'coordinate', None) # Changed V to K
                    # Присваиваем координату и добавляем буфер в результат
                    for item_in_buffer in active_items_buffer:
                        item_in_buffer["col_6_coord"] = price_total_coord_for_buffer
                    if first_section_found: # Доп. проверка, хотя должна быть True
                         processed_rows_list.extend(active_items_buffer)
                # else: Если цена 0 или пустая, буфер просто очистится ниже

                # Всегда очищаем буфер после строки "Всего по позиции"
                active_items_buffer = []

            elif row_type == "item":
                # first_section_found уже должен быть True, чтобы дойти сюда
                item_id_type = get_item_id_nature(cell_A.value)
                if item_id_type == "not_a_number":
                    continue
                
                # Добавляем проверку на пустоту колонки B
                cell_B = row_cells[1] if len(row_cells) > 1 else None
                if not cell_B or is_likely_empty(cell_B.value):
                    # print(f"[DEBUG] Пропуск строки {row_num}: Колонка B пуста.") # Debug print
                    continue # Пропускаем строку, если B пусто

                item_data = {
                    "type": "item",
                    "source_row_num": row_num,
                    "col_6_coord": None
                }
                # №, шифр, наименование, ед.изм, кол‑во  →  A, B, C, D, E
                input_indices_map = {1: 0, 2: 1, 3: 2, 4: 3, 5: 4}
                for out_col_num, in_col_idx in input_indices_map.items():
                    cell_to_map = row_cells[in_col_idx] if in_col_idx < len(row_cells) else None
                    item_data[f"col_{out_col_num}_coord"] = getattr(cell_to_map, 'coordinate', None)
                
                # --- Определение цены и решение о добавлении --- 
                price_coord_found = None
                price_is_range = False

                if item_id_type == "decimal": # Материал
                    cell_K_material = row_cells[item_total_cost_col_idx] if len(row_cells) > item_total_cost_col_idx else None
                    # Добавляем материал, только если цена не пустая и не 0
                    if cell_K_material and not is_zero_or_empty(cell_K_material.value):
                        price_coord_found = getattr(cell_K_material, 'coordinate', None)
                    # else: Материал с пустой/нулевой ценой не добавляется
                
                elif item_id_type == "integer": # Основная позиция
                    cell_K_inline_price = row_cells[item_total_cost_col_idx] if len(row_cells) > item_total_cost_col_idx else None
                    merge_KL_coord = None # check_merge(worksheet, row_num, item_total_cost_col_idx, item_total_cost_col_idx + 1)

                    if merge_KL_coord: # Если цена в объединенной ячейке K-L
                         # Получаем значение из верхней левой ячейки merge диапазона
                         value_to_check = getattr(cell_K_inline_price, 'value', None)
                         if not is_zero_or_empty(value_to_check):
                             price_coord_found = merge_KL_coord
                             price_is_range = True
                    elif cell_K_inline_price and not is_zero_or_empty(cell_K_inline_price.value): # Если есть значение в K (inline) и оно не 0
                        price_coord_found = getattr(cell_K_inline_price, 'coordinate', None)
                    
                    # Если цена не найдена inline (или =0), item пойдет в буфер
                    if price_coord_found is None:
                        active_items_buffer.append(item_data)
                        continue # Переходим к следующей строке, не добавляя item сразу

                # --- Добавление item в список (если цена найдена и != 0) ---
                if price_coord_found:
                    item_data["col_6_coord"] = price_coord_found
                    if price_is_range:
                         item_data["col_6_coord_is_range"] = True
                    processed_rows_list.append(item_data)
                # Если price_coord_found is None (для decimal цена была 0/пусто, 
                # для integer - item ушел в буфер выше), то item сюда не доходит или не добавляется

        # Сортируем все собранные строки по их исходному номеру строки
        processed_rows_list.sort(key=lambda x: x.get('source_row_num', float('inf')))
        
        all_coords_data = []
        for row_data in processed_rows_list:
            item_type = row_data.get("type")

            if item_type == "header_name":
                # ➊ координата A‑ячейки (левый‑верх диапазона A‑K)
                a_coord = get_start_coord(row_data.get("A_K_merge_coord"))  # 'A34'
                # output_headers должен быть доступен здесь, если нет, нужно его передать или определить
                # В текущем коде output_headers определен в начале функции process_grandsmeta_mixed
                header_row = [None] * len(output_headers)
                header_row[0] = a_coord                 # → колонка A
                header_row[1] = row_data.get("name_text")   # → колонка B
                # header_row[2] остаётся None по умолчанию из [None] * len(output_headers)
                all_coords_data.append(header_row)

            elif item_type == "header_footer":
                all_coords_data.append(['__FOOTER__',
                                        row_data.get("footer_text_content"),
                                        row_data.get("total_K_coord")]) # total_K_coord уже должен быть одиночной координатой

            elif item_type == "item":
                coords_row = [None] * len(output_headers) # Для item нужен этот массив
                for i_col in range(5): # Колонки 1-5 (индексы 0-4)
                    coords_row[i_col] = get_start_coord(row_data.get(f"col_{i_col+1}_coord"))
                coords_row[5] = get_start_coord(row_data.get("col_6_coord")) # Колонка 6 (ВСЕГО)
                all_coords_data.append(coords_row)

        return output_headers, all_coords_data

    except FileNotFoundError:
        return None, None
    except Exception as e:
        print(f"[КРИТИЧЕСКАЯ ОШИБКА] при обработке файла '{input_path}' (Турбосметчик-1, НОВЫЕ ПРАВИЛА v3): {e}")
        print("-" * 60); traceback.print_exc(); print("-" * 60)
        return None, None
    finally:
        if workbook:
            try: workbook.close()
            except Exception: pass