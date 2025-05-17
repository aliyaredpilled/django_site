# handlers/turbosmetchik/handler_v2.py
import openpyxl
import traceback
# Используем АБСОЛЮТНЫЙ импорт utils
from utils import is_likely_empty, check_merge, get_start_coord, get_item_id_nature # Заменили is_integer_like

def process_turbosmetchik_2(input_path):
    """
    ОБРАБАТЫВАЕТ один Excel файл по логике "Турбосметчик-2".
    Отличается от v1 маппингом 5-й колонки выхода (Кол-во) на колонку N входа.
    ВОЗВРАЩАЕТ данные (заголовки и координаты) для дальнейшей обработки.
    """
    # print(f"\n--- Обработка файла (Турбосметчик-2): {os.path.basename(input_path)} ---")

    output_headers = ["№№ п/п", "Шифр расценки и коды ресурсов", "Наименование работ и затрат", "Единица измерения", "Кол-во единиц", "ВСЕГО затрат, руб."]
    start_id_col_idx = 0    # A
    # Для Турбосметчик-2 (аналогично v1 по логике цен):
    # - Цена для "Всего по позиции" (для item'ов в буфере) берется из колонки V.
    # - Цена для item'а с inline-ценой (merge V-W) берется из V.
    # - Цена для МАТЕРИАЛА (дробный номер) теперь тоже будет браться из V его строки.
    item_total_cost_col_idx = 21 # V - колонка с итоговой ценой

    processed_rows_list = []
    active_items_buffer = [] # Буфер для основных позиций (целый номер), ожидающих общую цену
    pending_section_header = None
    pending_subsection_header = None
    first_section_found = False

    workbook = None
    try:
        workbook = openpyxl.load_workbook(filename=input_path, data_only=True)
        if not workbook.sheetnames:
            # print(f"Ошибка: Нет листов в файле '{input_path}'.") # Убираем print для чистоты
            return None, None
        worksheet = workbook[workbook.sheetnames[0]]

        for row_num, row_cells_tuple in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
            row_cells = list(row_cells_tuple)
            non_empty_info = [(i, getattr(c, 'coordinate', None), c.value) for i, c in enumerate(row_cells) if not is_likely_empty(c.value)]
            if not non_empty_info:
                continue

            cell_A = row_cells[start_id_col_idx] if len(row_cells) > start_id_col_idx else None
            cell_A_value_str = str(getattr(cell_A, 'value', '')).strip() if cell_A else ""

            row_type = None
            header_merge_coord = check_merge(worksheet, row_num, 0, 22) # A(0) - W(22)
            footer_merge_coord_DK = check_merge(worksheet, row_num, 3, 10) # D(3) - K(10)

            cell_D = row_cells[3] if len(row_cells) > 3 else None
            cell_D_value_str = str(getattr(cell_D, 'value', '')).strip() if cell_D else ""

            if header_merge_coord and cell_A_value_str.startswith("Раздел"):
                row_type = "section_header"
            elif header_merge_coord and cell_A_value_str.startswith("Подраздел"):
                row_type = "subsection_header"
            elif footer_merge_coord_DK and cell_D_value_str.startswith("Итого по подразделу"):
                row_type = "subsection_footer"
            elif footer_merge_coord_DK and cell_D_value_str.startswith("Итого по разделу"):
                 row_type = "section_footer"
            else:
                dr_merge_coord = check_merge(worksheet, row_num, 3, 17) # D(3) - R(17)
                cell_D_price_text_obj = row_cells[3] if len(row_cells) > 3 else None
                if dr_merge_coord and cell_D_price_text_obj and str(getattr(cell_D_price_text_obj, 'value', '')).strip() == "Всего по позиции":
                    row_type = "item_price_row"
                else:
                    if cell_A and cell_A.data_type != 'f' and not is_likely_empty(cell_A.value):
                         try:
                             float(str(cell_A.value).replace(',', '.').strip()) # Предварительная проверка на число
                             row_type = "item"
                         except (ValueError, TypeError):
                             pass

            if row_type in ["section_header", "subsection_header", "section_footer", "subsection_footer"] and active_items_buffer:
                if first_section_found:
                    processed_rows_list.extend(active_items_buffer)
                active_items_buffer = []

            if row_type == "section_header":
                if first_section_found:
                    if pending_subsection_header: processed_rows_list.append(pending_subsection_header)
                    if pending_section_header: processed_rows_list.append(pending_section_header)
                pending_section_header = {"type": "header", "level": "section", "start_row": row_num, "col_1_coord": header_merge_coord, "col_3_value": cell_A_value_str, "col_6_coord": None}
                pending_subsection_header = None
                first_section_found = True
            elif row_type == "subsection_header":
                if first_section_found and pending_subsection_header: processed_rows_list.append(pending_subsection_header)
                pending_subsection_header = {"type": "header", "level": "subsection", "start_row": row_num, "col_1_coord": header_merge_coord, "col_3_value": cell_A_value_str, "col_6_coord": None}
            elif row_type == "subsection_footer":
                 if pending_subsection_header:
                    cell_V_footer = row_cells[item_total_cost_col_idx] if len(row_cells) > item_total_cost_col_idx else None
                    pending_subsection_header["col_6_coord"] = getattr(cell_V_footer, 'coordinate', None)
                    if first_section_found:
                        processed_rows_list.append(pending_subsection_header)
                        pending_subsection_header = None
            elif row_type == "section_footer":
                if first_section_found and pending_subsection_header:
                    processed_rows_list.append(pending_subsection_header)
                    pending_subsection_header = None
                if pending_section_header:
                    cell_V_footer = row_cells[item_total_cost_col_idx] if len(row_cells) > item_total_cost_col_idx else None
                    pending_section_header["col_6_coord"] = getattr(cell_V_footer, 'coordinate', None)
                    if first_section_found:
                        processed_rows_list.append(pending_section_header)
                        pending_section_header = None
            elif row_type == "item_price_row": # "Всего по позиции"
                cell_V_price = row_cells[item_total_cost_col_idx] if len(row_cells) > item_total_cost_col_idx else None
                price_total_coord_for_buffer = getattr(cell_V_price, 'coordinate', None)
                for item in active_items_buffer: # Применяем к элементам в буфере
                    item["col_6_coord"] = price_total_coord_for_buffer
                if first_section_found and active_items_buffer:
                     processed_rows_list.extend(active_items_buffer)
                     active_items_buffer = []
            elif row_type == "item":
                if first_section_found and cell_A: # Убедимся, что cell_A существует
                    item_id_type = get_item_id_nature(cell_A.value)
                    
                    if item_id_type == "not_a_number": # Если это не числовой ID, пропускаем
                        continue

                    item_data = {"type": "item", "start_row": row_num, "col_6_coord": None}
                    # Вход T2: A(0)   B(1)    D(3)           L(11)       N(13) <-- ОТЛИЧИЕ от T1
                    input_indices_map = {1: 0, 2: 1, 3: 3, 4: 11, 5: 13} # Колонка "Кол-во" теперь N(13)

                    for out_col_num, in_col_idx in input_indices_map.items():
                        cell_to_map = row_cells[in_col_idx] if in_col_idx < len(row_cells) else None
                        item_data[f"col_{out_col_num}_coord"] = getattr(cell_to_map, 'coordinate', None)
                    
                    if item_id_type == "decimal": # Это МАТЕРИАЛ
                        # Цена для материала берется из колонки V (item_total_cost_col_idx) ТЕКУЩЕЙ строки
                        cell_V_material = row_cells[item_total_cost_col_idx] if len(row_cells) > item_total_cost_col_idx else None
                        item_data["col_6_coord"] = getattr(cell_V_material, 'coordinate', None)
                        processed_rows_list.append(item_data) # Материалы добавляются сразу

                    elif item_id_type == "integer": # Это ОСНОВНАЯ ПОЗИЦИЯ
                        inline_price_coord = None
                        # Проверяем наличие цены в этой же строке (merge V-W)
                        # V - item_total_cost_col_idx (21)
                        # W - item_total_cost_col_idx + 1 (22)
                        merge_VW_coord = check_merge(worksheet, row_num, item_total_cost_col_idx, item_total_cost_col_idx + 1)
                        if merge_VW_coord:
                            cell_V_inline = row_cells[item_total_cost_col_idx] if len(row_cells) > item_total_cost_col_idx else None
                            if cell_V_inline and not is_likely_empty(cell_V_inline.value):
                                inline_price_coord = get_start_coord(merge_VW_coord) # Берем начало merge
                        
                        if inline_price_coord:
                            # Если цена найдена в строке (inline), присваиваем и добавляем
                            item_data["col_6_coord"] = inline_price_coord
                            processed_rows_list.append(item_data)
                        else:
                            # Если цена не найдена (или номер не целый, но это уже обработано выше), добавляем item в буфер
                            active_items_buffer.append(item_data)

        if first_section_found:
            if active_items_buffer: processed_rows_list.extend(active_items_buffer)
            if pending_subsection_header: processed_rows_list.append(pending_subsection_header)
            if pending_section_header: processed_rows_list.append(pending_section_header)

        processed_rows_list.sort(key=lambda x: x.get('start_row', float('inf')))
        all_coords_data = []
        for row_data in processed_rows_list:
            coords_row = [None] * len(output_headers)
            item_type = row_data.get("type")

            if item_type == "header":
                coords_row[0] = get_start_coord(row_data.get("col_1_coord"))
                coords_row[2] = row_data.get("col_3_value")
                coords_row[5] = get_start_coord(row_data.get("col_6_coord"))
            elif item_type == "item":
                for i_col in range(5): # Колонки 1-5
                    coords_row[i_col] = get_start_coord(row_data.get(f"col_{i_col+1}_coord"))
                coords_row[5] = get_start_coord(row_data.get("col_6_coord")) # Колонка 6 (ВСЕГО)

            all_coords_data.append(coords_row)

        return output_headers, all_coords_data

    except FileNotFoundError:
        # print(f"[ОШИБКА] Файл не найден: {input_path}") # Убираем print для чистоты
        return None, None
    except Exception as e:
        print(f"[КРИТИЧЕСКАЯ ОШИБКА] при обработке файла '{input_path}' (Турбосметчик-2): {e}")
        print("-" * 60); traceback.print_exc(); print("-" * 60)
        return None, None
    finally:
        if workbook:
            try: workbook.close()
            except Exception: pass # Игнорируем ошибки закрытия