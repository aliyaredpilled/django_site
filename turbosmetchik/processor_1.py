import openpyxl
import traceback
# Используем АБСОЛЮТНЫЙ импорт utils
from utils import is_likely_empty, check_merge, get_start_coord, get_item_id_nature

def process_turbosmetchik_1(input_path):
    """
    ОБРАБАТЫВАЕТ один Excel файл по НОВЫМ правилам "Турбосметчик-1".
    ВОЗВРАЩАЕТ данные (заголовки и координаты) для дальнейшей обработки.
    """
    output_headers = ["№№ п/п", "Шифр расценки и коды ресурсов", "Наименование работ и затрат", "Единица измерения", "Кол-во единиц", "ВСЕГО затрат, руб."]
    start_id_col_idx = 0    # A
    # Для Турбосметчик-1 (цены, вероятно, остаются теми же):
    # - Цена для "Всего по позиции" (для item'ов в буфере) берется из колонки V.
    # - Цена для item'а с inline-ценой (merge V-W) берется из V.
    # - Цена для МАТЕРИАЛА (дробный номер) теперь тоже будет браться из V его строки.
    item_total_cost_col_idx = 21 # V - колонка с итоговой ценой

    processed_rows_list = []
    active_items_buffer = [] # Буфер для основных позиций (целый номер), ожидающих общую цену
    pending_section_header = None
    pending_subsection_header = None
    first_section_found = False

    workbook = None
    try:
        workbook = openpyxl.load_workbook(filename=input_path, data_only=True)
        if not workbook.sheetnames:
            return None, None
        worksheet = workbook[workbook.sheetnames[0]]

        for row_num, row_cells_tuple in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
            row_cells = list(row_cells_tuple)
            non_empty_info = [(i, getattr(c, 'coordinate', None), c.value) for i, c in enumerate(row_cells) if not is_likely_empty(c.value)]
            if not non_empty_info:
                continue

            cell_A = row_cells[start_id_col_idx] if len(row_cells) > start_id_col_idx else None
            cell_A_value_str = str(getattr(cell_A, 'value', '')).strip() if cell_A else ""

            # Получаем значение из ячейки C для правил футеров и "Всего по позиции"
            cell_C = row_cells[2] if len(row_cells) > 2 else None # C - индекс 2
            cell_C_value_str = str(getattr(cell_C, 'value', '')).strip() if cell_C else ""

            row_type = None
            # НОВЫЕ ПРАВИЛА для определения типа строки:
            
            # 1. Проверка на заголовки (Раздел/Подраздел)
            # Ячейки с A по K объединены (индексы 0-10)
            header_merge_AK_coord = check_merge(worksheet, row_num, 0, 10) # A(0) - K(10)

            if header_merge_AK_coord:
                if cell_A_value_str.startswith("Раздел"):
                    row_type = "section_header"
                else: # Если A-K объединены и не начинается с "Раздел", то это подраздел
                    row_type = "subsection_header"
            else:
                # 2. Проверка на футеры (Итоги)
                # Ячейки с C по H объединены (индексы 2-7)
                footer_merge_CH_coord = check_merge(worksheet, row_num, 2, 7) # C(2) - H(7)
                if footer_merge_CH_coord:
                    if cell_C_value_str.startswith("Итого по разделу"):
                        row_type = "section_footer"
                    elif cell_C_value_str.startswith("Итого по подразделу"):
                        row_type = "subsection_footer"
                # 3. Проверка на "Всего по позиции" (если не футер)
                elif cell_C_value_str == "Всего по позиции":
                     # По новым правилам, для "Всего по позиции" не указано объединение ячеек,
                     # только значение в ячейке C.
                    row_type = "item_price_row"
                else:
                    # 4. Проверка на позицию (item)
                    if cell_A and cell_A.data_type != 'f' and not is_likely_empty(cell_A.value):
                         try:
                             float(str(cell_A.value).replace(',', '.').strip()) # Предварительная проверка на число
                             row_type = "item"
                         except (ValueError, TypeError):
                             pass # Не число, значит не item по этому правилу

            # --- Остальная логика обработки буферов и добавления данных ---
            # Эта часть в основном остается прежней, но использует новые row_type

            if row_type in ["section_header", "subsection_header", "section_footer", "subsection_footer"] and active_items_buffer:
                if first_section_found:
                    processed_rows_list.extend(active_items_buffer)
                active_items_buffer = []

            if row_type == "section_header":
                if first_section_found:
                    if pending_subsection_header: processed_rows_list.append(pending_subsection_header)
                    if pending_section_header: processed_rows_list.append(pending_section_header)
                # Используем header_merge_AK_coord для col_1_coord
                pending_section_header = {"type": "header", "level": "section", "start_row": row_num, "col_1_coord": header_merge_AK_coord, "col_3_value": cell_A_value_str, "col_6_coord": None}
                pending_subsection_header = None
                first_section_found = True
            elif row_type == "subsection_header":
                if first_section_found and pending_subsection_header: processed_rows_list.append(pending_subsection_header)
                # Используем header_merge_AK_coord для col_1_coord
                pending_subsection_header = {"type": "header", "level": "subsection", "start_row": row_num, "col_1_coord": header_merge_AK_coord, "col_3_value": cell_A_value_str, "col_6_coord": None}
            elif row_type == "subsection_footer":
                 if pending_subsection_header:
                    cell_V_footer = row_cells[item_total_cost_col_idx] if len(row_cells) > item_total_cost_col_idx else None
                    pending_subsection_header["col_6_coord"] = getattr(cell_V_footer, 'coordinate', None)
                    if first_section_found:
                        processed_rows_list.append(pending_subsection_header)
                        pending_subsection_header = None
            elif row_type == "section_footer":
                if first_section_found and pending_subsection_header:
                    processed_rows_list.append(pending_subsection_header)
                    pending_subsection_header = None
                if pending_section_header:
                    cell_V_footer = row_cells[item_total_cost_col_idx] if len(row_cells) > item_total_cost_col_idx else None
                    pending_section_header["col_6_coord"] = getattr(cell_V_footer, 'coordinate', None)
                    if first_section_found:
                        processed_rows_list.append(pending_section_header)
                        pending_section_header = None
            elif row_type == "item_price_row": # "Всего по позиции"
                cell_V_price = row_cells[item_total_cost_col_idx] if len(row_cells) > item_total_cost_col_idx else None
                price_total_coord_for_buffer = getattr(cell_V_price, 'coordinate', None)
                for item in active_items_buffer: # Применяем к элементам в буфере
                    item["col_6_coord"] = price_total_coord_for_buffer
                if first_section_found and active_items_buffer:
                     processed_rows_list.extend(active_items_buffer)
                     active_items_buffer = []
            elif row_type == "item":
                if first_section_found and cell_A:
                    item_id_type = get_item_id_nature(cell_A.value)
                    
                    if item_id_type == "not_a_number":
                        continue

                    item_data = {"type": "item", "start_row": row_num, "col_6_coord": None}
                    # Маппинг колонок для item остается тем же
                    # Вход T1: A(0)   B(1)    D(3)           L(11)       M(12)
                    input_indices_map = {1: 0, 2: 1, 3: 3, 4: 11, 5: 12}

                    for out_col_num, in_col_idx in input_indices_map.items():
                        cell_to_map = row_cells[in_col_idx] if in_col_idx < len(row_cells) else None
                        item_data[f"col_{out_col_num}_coord"] = getattr(cell_to_map, 'coordinate', None)
                    
                    if item_id_type == "decimal": # Это МАТЕРИАЛ
                        cell_V_material = row_cells[item_total_cost_col_idx] if len(row_cells) > item_total_cost_col_idx else None
                        item_data["col_6_coord"] = getattr(cell_V_material, 'coordinate', None)
                        processed_rows_list.append(item_data)

                    elif item_id_type == "integer": # Это ОСНОВНАЯ ПОЗИЦИЯ
                        inline_price_coord = None
                        # Проверка на inline-цену (merge V-W) остается той же
                        merge_VW_coord = check_merge(worksheet, row_num, item_total_cost_col_idx, item_total_cost_col_idx + 1)
                        if merge_VW_coord:
                            cell_V_inline = row_cells[item_total_cost_col_idx] if len(row_cells) > item_total_cost_col_idx else None
                            if cell_V_inline and not is_likely_empty(cell_V_inline.value):
                                inline_price_coord = get_start_coord(merge_VW_coord)
                        
                        if inline_price_coord:
                            item_data["col_6_coord"] = inline_price_coord
                            processed_rows_list.append(item_data)
                        else:
                            active_items_buffer.append(item_data)

        # --- Финальная обработка и возврат данных ---
        # Эта часть также остается в основном без изменений

        if first_section_found:
            if active_items_buffer: processed_rows_list.extend(active_items_buffer)
            if pending_subsection_header: processed_rows_list.append(pending_subsection_header)
            if pending_section_header: processed_rows_list.append(pending_section_header)

        processed_rows_list.sort(key=lambda x: x.get('start_row', float('inf')))
        all_coords_data = []
        for row_data in processed_rows_list:
            coords_row = [None] * len(output_headers)
            item_type = row_data.get("type")

            if item_type == "header":
                coords_row[0] = get_start_coord(row_data.get("col_1_coord"))
                coords_row[2] = row_data.get("col_3_value")
                coords_row[5] = get_start_coord(row_data.get("col_6_coord"))
            elif item_type == "item":
                for i_col in range(5): # Колонки 1-5
                    coords_row[i_col] = get_start_coord(row_data.get(f"col_{i_col+1}_coord"))
                coords_row[5] = get_start_coord(row_data.get("col_6_coord")) # Колонка 6 (ВСЕГО)

            all_coords_data.append(coords_row)

        return output_headers, all_coords_data

    except FileNotFoundError:
        return None, None
    except Exception as e:
        print(f"[КРИТИЧЕСКАЯ ОШИБКА] при обработке файла '{input_path}' (Турбосметчик-1, НОВЫЕ ПРАВИЛА): {e}")
        print("-" * 60); traceback.print_exc(); print("-" * 60)
        return None, None
    finally:
        if workbook:
            try: workbook.close()
            except Exception: pass