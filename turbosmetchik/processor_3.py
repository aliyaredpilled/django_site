# handlers/turbosmetchik/processor_3.py
import openpyxl
import traceback
# Используем АБСОЛЮТНЫЙ импорт utils
from utils import is_likely_empty, check_merge, get_start_coord, get_item_id_nature # Заменили is_integer_like

def process_turbosmetchik_3(input_path):
    """
    ОБРАБАТЫВАЕТ один Excel файл по логике "Турбосметчик-3" (Новый Формат Смет).
    ВОЗВРАЩАЕТ данные (заголовки и координаты) для дальнейшей обработки.
    """
    # print(f"\n--- Обработка файла (Турбосметчик-3): {os.path.basename(input_path)} ---")

    output_headers = ["№№ п/п", "Шифр расценки и коды ресурсов", "Наименование работ и затрат", "Единица измерения", "Кол-во единиц", "ВСЕГО затрат, руб."]
    start_id_col_idx = 0    # A
    # Для Турбосметчик-3 (Новый Формат):
    # - Итоговая цена для футеров и "Всего по позиции" (для буфера) берется из колонки Z.
    # - Inline-цена для item'а (если есть, merge Z-AB) берется из Z.
    # - Цена для МАТЕРИАЛА (дробный номер) теперь тоже будет браться из Z его строки.
    item_total_cost_col_idx = 25 # Z - колонка с итоговой ценой

    processed_rows_list = []
    active_items_buffer = [] # Буфер для основных позиций, ожидающих "Всего по позиции"
    pending_section_header = None
    pending_subsection_header = None
    first_section_found = False

    workbook = None
    try:
        workbook = openpyxl.load_workbook(filename=input_path, data_only=True)
        if not workbook.sheetnames:
            # print(f"Ошибка: Нет листов в файле '{input_path}'.") # Убираем print для чистоты
            return None, None
        worksheet = workbook[workbook.sheetnames[0]]

        max_col_idx = worksheet.max_column - 1 # 0-based index

        for row_num, row_cells_tuple in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
            row_cells = list(row_cells_tuple)
            non_empty_info = [(i, getattr(c, 'coordinate', None), c.value) for i, c in enumerate(row_cells) if not is_likely_empty(c.value)]
            if not non_empty_info:
                continue

            cell_A = row_cells[start_id_col_idx] if start_id_col_idx <= max_col_idx else None
            cell_A_value_str = str(getattr(cell_A, 'value', '')).strip() if cell_A else ""

            row_type = None
            header_merge_coord_A_AB = check_merge(worksheet, row_num, 0, 27) # A-AB
            footer_merge_coord_E_I = check_merge(worksheet, row_num, 4, 8)  # E-I
            
            cell_E_idx = 4
            cell_E = row_cells[cell_E_idx] if cell_E_idx <= max_col_idx else None
            cell_E_value_str = str(getattr(cell_E, 'value', '')).strip() if cell_E else ""

            if header_merge_coord_A_AB and cell_A_value_str.startswith("Раздел"):
                row_type = "section_header"
            elif header_merge_coord_A_AB and cell_A_value_str.startswith("Подраздел"):
                row_type = "subsection_header"
            elif footer_merge_coord_E_I and cell_E_value_str.startswith("Итого по подразделу"):
                row_type = "subsection_footer"
            elif footer_merge_coord_E_I and cell_E_value_str.startswith("Итого по разделу"):
                row_type = "section_footer"
            else:
                item_price_merge_coord_E_S = check_merge(worksheet, row_num, 4, 18) # E-S
                # cell_E_price_text используется из уже полученной cell_E
                if item_price_merge_coord_E_S and cell_E and str(getattr(cell_E, 'value', '')).strip() == "Всего по позиции":
                    row_type = "item_price_row"
                else:
                    if cell_A and cell_A.data_type != 'f' and not is_likely_empty(cell_A.value):
                         try:
                             float(str(cell_A.value).replace(',', '.').strip()) # Предварительная проверка на число
                             row_type = "item"
                         except (ValueError, TypeError):
                             pass

            if row_type in ["section_header", "subsection_header", "section_footer", "subsection_footer"] and active_items_buffer:
                if first_section_found:
                    processed_rows_list.extend(active_items_buffer)
                active_items_buffer = []

            if row_type == "section_header":
                if first_section_found:
                    if pending_subsection_header: processed_rows_list.append(pending_subsection_header)
                    if pending_section_header: processed_rows_list.append(pending_section_header)
                pending_section_header = {"type": "header", "level": "section", "start_row": row_num, "col_1_coord": header_merge_coord_A_AB, "col_3_value": cell_A_value_str, "col_6_coord": None}
                pending_subsection_header = None
                first_section_found = True
            elif row_type == "subsection_header":
                if first_section_found and pending_subsection_header: processed_rows_list.append(pending_subsection_header)
                pending_subsection_header = {"type": "header", "level": "subsection", "start_row": row_num, "col_1_coord": header_merge_coord_A_AB, "col_3_value": cell_A_value_str, "col_6_coord": None}
            elif row_type == "subsection_footer":
                 if pending_subsection_header:
                    cell_Z_footer = row_cells[item_total_cost_col_idx] if item_total_cost_col_idx <= max_col_idx else None
                    pending_subsection_header["col_6_coord"] = getattr(cell_Z_footer, 'coordinate', None)
                    if first_section_found:
                        processed_rows_list.append(pending_subsection_header)
                        pending_subsection_header = None
            elif row_type == "section_footer":
                if first_section_found and pending_subsection_header:
                    processed_rows_list.append(pending_subsection_header)
                    pending_subsection_header = None
                if pending_section_header:
                    cell_Z_footer = row_cells[item_total_cost_col_idx] if item_total_cost_col_idx <= max_col_idx else None
                    pending_section_header["col_6_coord"] = getattr(cell_Z_footer, 'coordinate', None)
                    if first_section_found:
                        processed_rows_list.append(pending_section_header)
                        pending_section_header = None
            elif row_type == "item_price_row": # "Всего по позиции"
                price_total_coord_for_buffer = None
                cell_Z_price = row_cells[item_total_cost_col_idx] if item_total_cost_col_idx <= max_col_idx else None
                if cell_Z_price:
                    # Проверка объединения Z-AB для строки "Всего по позиции"
                    merge_ZAB_coord = check_merge(worksheet, row_num, item_total_cost_col_idx, item_total_cost_col_idx + 2) # Z(25)-AB(27)
                    if merge_ZAB_coord:
                        price_total_coord_for_buffer = get_start_coord(merge_ZAB_coord)
                    else:
                        price_total_coord_for_buffer = getattr(cell_Z_price, 'coordinate', None) # Если не merge, берем просто Z

                for item in active_items_buffer:
                    item["col_6_coord"] = price_total_coord_for_buffer
                if first_section_found and active_items_buffer:
                     processed_rows_list.extend(active_items_buffer)
                     active_items_buffer = []
            elif row_type == "item":
                if first_section_found and cell_A:
                    item_id_type = get_item_id_nature(cell_A.value)

                    if item_id_type == "not_a_number":
                        continue
                        
                    item_data = {"type": "item", "start_row": row_num, "col_6_coord": None}
                    # Вход NF: A(0)   B(1)    E(4)           J(9)       M(12)
                    input_indices_map = {1: 0, 2: 1, 3: 4, 4: 9, 5: 12}

                    for out_col_num, in_col_idx in input_indices_map.items():
                        cell_to_map = row_cells[in_col_idx] if in_col_idx <= max_col_idx else None
                        item_data[f"col_{out_col_num}_coord"] = getattr(cell_to_map, 'coordinate', None)
                    
                    if item_id_type == "decimal": # Это МАТЕРИАЛ
                        # Цена для материала берется из колонки Z (item_total_cost_col_idx) ТЕКУЩЕЙ строки
                        # Для материалов не ожидается merge Z-AB, просто значение из Z
                        cell_Z_material = row_cells[item_total_cost_col_idx] if item_total_cost_col_idx <= max_col_idx else None
                        item_data["col_6_coord"] = getattr(cell_Z_material, 'coordinate', None)
                        processed_rows_list.append(item_data)

                    elif item_id_type == "integer": # Это ОСНОВНАЯ ПОЗИЦИЯ
                        inline_price_coord = None
                        # Проверяем наличие цены в этой же строке (merge Z-AB)
                        merge_ZAB_coord_inline = check_merge(worksheet, row_num, item_total_cost_col_idx, item_total_cost_col_idx + 2) # Z(25)-AB(27)
                        if merge_ZAB_coord_inline:
                            cell_Z_inline = row_cells[item_total_cost_col_idx] if item_total_cost_col_idx <= max_col_idx else None
                            if cell_Z_inline and not is_likely_empty(cell_Z_inline.value):
                                inline_price_coord = get_start_coord(merge_ZAB_coord_inline)
                        
                        if inline_price_coord:
                            item_data["col_6_coord"] = inline_price_coord
                            processed_rows_list.append(item_data)
                        else:
                            active_items_buffer.append(item_data)

        if first_section_found:
            if active_items_buffer: processed_rows_list.extend(active_items_buffer)
            if pending_subsection_header: processed_rows_list.append(pending_subsection_header)
            if pending_section_header: processed_rows_list.append(pending_section_header)

        processed_rows_list.sort(key=lambda x: x.get('start_row', float('inf')))
        all_coords_data = []
        for row_data in processed_rows_list:
            coords_row = [None] * len(output_headers)
            item_type = row_data.get("type")

            if item_type == "header":
                coords_row[0] = get_start_coord(row_data.get("col_1_coord"))
                coords_row[2] = row_data.get("col_3_value")
                coords_row[5] = get_start_coord(row_data.get("col_6_coord"))
            elif item_type == "item":
                for i_col in range(5):
                    coords_row[i_col] = get_start_coord(row_data.get(f"col_{i_col+1}_coord"))
                coords_row[5] = get_start_coord(row_data.get("col_6_coord"))

            all_coords_data.append(coords_row)

        return output_headers, all_coords_data

    except FileNotFoundError:
        # print(f"[ОШИБКА] Файл не найден: {input_path}") # Убираем print для чистоты
        return None, None
    except Exception as e:
        print(f"[КРИТИЧЕСКАЯ ОШИБКА] при обработке файла '{input_path}' (Турбосметчик-3): {e}")
        print("-" * 60); traceback.print_exc(); print("-" * 60)
        return None, None
    finally:
        if workbook:
            try: workbook.close()
            except Exception: pass