from __future__ import annotations

import logging
import traceback
import re 
from enum import Enum, auto
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl # type: ignore

from utils import (
    check_merge,
    get_item_id_nature,
    get_start_coord,
    is_likely_empty,
    is_zero,
)

logger = logging.getLogger(__name__)
# --------------------------------------------------------------------------- #
#                                Константы                                    #
# --------------------------------------------------------------------------- #
OUTPUT_HEADERS = [
    "№№ п/п", "Шифр расценки и коды ресурсов", "Наименование работ и затрат",
    "Единица измерения", "Кол-во единиц", "ВСЕГО затрат, руб.",
]
KEY_TOTAL_SMALL_CASE = "Всего по позиции"
KEY_TOTAL_WORK_UPPER_CASE = "ВСЕГО работ по позиции"
KEY_TOTAL_ALL_UPPER_CASE = "ВСЕГО по позиции"
HEADER_SECTION = "Раздел"
HEADER_SUBSECTION = "Подраздел"
FOOTER_LOCAL_ESTIMATE_PREFIX = "Итого по локальной смете"
FOOTER_SECTION = "Итого по разделу"
FOOTER_SUBSECTION = "Итого по подразделу"
LOCAL_ESTIMATE_HEADER_PREFIX_TO_STRIP_LOWER = "локальная смета"

class SmetaType(Enum):
    CLASSIC = auto()
    TOTAL_SMALL_CASE = auto()
    TOTAL_UPPER_CASE = auto()

# --------------------------------------------------------------------------- #
#                       Вспомогательные функции                               #
# --------------------------------------------------------------------------- #
def _is_column_number_row(row_cells: Tuple[Any, ...], max_check: int = 11) -> bool:
    for idx in range(max_check):
        if idx >= len(row_cells) or str(row_cells[idx].value).strip() != str(idx + 1): return False
    return all(is_likely_empty(cell.value) for cell in row_cells[max_check:])

def _determine_smeta_type(ws: openpyxl.worksheet.worksheet.Worksheet) -> SmetaType:
    logger.info("Определяю тип сметы…")
    saw_small, saw_work_upper = False, False
    for row_idx in range(2, ws.max_row + 1):
        cell_c_obj = ws.cell(row=row_idx, column=3)
        val = str(cell_c_obj.value).strip() if cell_c_obj and not is_likely_empty(cell_c_obj.value) else ""
        if not val: continue
        if val.startswith(KEY_TOTAL_ALL_UPPER_CASE): logger.info("Тип:TOTAL_UPPER_CASE(по'%s')",KEY_TOTAL_ALL_UPPER_CASE);return SmetaType.TOTAL_UPPER_CASE
        if val.startswith(KEY_TOTAL_WORK_UPPER_CASE): saw_work_upper = True
        if val.startswith(KEY_TOTAL_SMALL_CASE): saw_small = True
    if saw_work_upper: logger.info("Тип:TOTAL_UPPER_CASE(по'%s')",KEY_TOTAL_WORK_UPPER_CASE);return SmetaType.TOTAL_UPPER_CASE
    if saw_small: logger.info("Тип:TOTAL_SMALL_CASE(по'%s')",KEY_TOTAL_SMALL_CASE);return SmetaType.TOTAL_SMALL_CASE
    logger.info("Тип сметы: CLASSIC"); return SmetaType.CLASSIC

def _normalize_header_name(text: str) -> str:
    text_lower = str(text).lower().strip()
    if text_lower.startswith(LOCAL_ESTIMATE_HEADER_PREFIX_TO_STRIP_LOWER):
        text_lower = text_lower[len(LOCAL_ESTIMATE_HEADER_PREFIX_TO_STRIP_LOWER):].strip()
        if text_lower.startswith(":"): text_lower = text_lower[1:].strip()
    text_lower = re.sub(r'["\'«»]', '', text_lower); text_lower = re.sub(r'\s+', ' ', text_lower).strip()
    return text_lower

def _extract_and_normalize_footer_name(full_footer_text: str) -> str:
    norm_full_footer = _normalize_header_name(full_footer_text)
    norm_footer_prefix = _normalize_header_name(FOOTER_LOCAL_ESTIMATE_PREFIX)
    if norm_full_footer.startswith(norm_footer_prefix):
        name_part = norm_full_footer[len(norm_footer_prefix):].strip()
        if name_part.startswith(":"): name_part = name_part[1:].strip()
        return name_part
    return ""

# --------------------------------------------------------------------------- #
#                             Основная логика                                 #
# --------------------------------------------------------------------------- #
def process_smeta_ru(path: str | Path) -> Tuple[Optional[List[str]], Optional[List[List[Optional[str]]]]]:
    path_obj = Path(path)
    wb = None
    try:
        wb = openpyxl.load_workbook(path_obj, data_only=True)
        if not wb.sheetnames: logger.error("Файл %s нет листов.", path_obj); return None, None
        ws: openpyxl.worksheet.worksheet.Worksheet = wb[wb.sheetnames[0]]
        smeta_type = _determine_smeta_type(ws)

        material_col_idx, item_total_price_value_col_idx, footer_text_end_col_idx, \
        footer_price_start_col_idx, footer_price_end_col_idx, footer_price_value_col_idx = \
            (9, 8, 7, 8, 9, 8) if smeta_type is SmetaType.CLASSIC else (10, 9, 6, 9, 10, 9)

        processed_rows: List[Dict[str, Any]] = []
        active_items: List[Dict[str, Any]] = []
        
        potential_local_estimate_candidates: List[Dict[str, Any]] = []
        pending_section: Optional[Dict[str, Any]] = None
        pending_subsection: Optional[Dict[str, Any]] = None
        
        found_upper_prices: Dict[str, Dict[str, Any]] = {}
        skipped_zero_price_individual, skipped_zero_price_total = 0, 0

        for row_idx, row_cells_tuple in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            if _is_column_number_row(row_cells_tuple): continue
            current_row_cells: List[openpyxl.cell.cell.Cell] = list(row_cells_tuple)
            non_empty_cols_indices = {idx for idx,c in enumerate(current_row_cells) if not is_likely_empty(c.value)}
            if not non_empty_cols_indices: continue
            cell_a = current_row_cells[0] if len(current_row_cells)>0 else None
            cell_c = current_row_cells[2] if len(current_row_cells)>2 else None
            a_val = str(cell_a.value).strip() if cell_a and not is_likely_empty(cell_a.value) else ""
            c_val = str(cell_c.value).strip() if cell_c and not is_likely_empty(cell_c.value) else ""

            row_type: Optional[str] = None; item_nature: Optional[str] = None
            is_price_line: bool = False; current_price_line_info: Dict[str, Any] = {}
            footer_price_merged_coord: Optional[str] = None; header_ak_merge_coord: Optional[str] = None
            
            header_ak_merge_coord = check_merge(ws, row_idx, 0, 10)
            if header_ak_merge_coord:
                if a_val.startswith(HEADER_SECTION): row_type = "section_header"
                elif a_val.startswith(HEADER_SUBSECTION): row_type = "subsection_header"
                else: 
                    norm_name = _normalize_header_name(a_val)
                    cand_data = {"type":"header","level":"local_estimate_candidate", "start_row":row_idx,
                                 "col_1_coord":header_ak_merge_coord,"col_3_value":a_val,"normalized_name":norm_name,
                                 "col_6_value":None,"col_6_coord":None}
                    potential_local_estimate_candidates.append(cand_data)
                    logger.debug(f"Стр.{row_idx}:Добавл.канд.в ЛС:'{a_val}'(норм:'{norm_name}')")
            
            if row_type is None:
                txt_m = check_merge(ws,row_idx,0,footer_text_end_col_idx)
                prc_m = check_merge(ws,row_idx,footer_price_start_col_idx,footer_price_end_col_idx)
                if txt_m and prc_m:
                    if a_val.startswith(FOOTER_LOCAL_ESTIMATE_PREFIX):row_type="local_estimate_footer"
                    elif a_val.startswith(FOOTER_SECTION):row_type="section_footer"
                    elif a_val.startswith(FOOTER_SUBSECTION):row_type="subsection_footer"
                    if row_type and row_type.endswith("_footer"):footer_price_merged_coord=prc_m
            
            if row_type is None:
                if smeta_type is SmetaType.CLASSIC:
                    n_e={item_total_price_value_col_idx,10};e_b=item_total_price_value_col_idx
                    if n_e.issubset(non_empty_cols_indices) and all(i not in non_empty_cols_indices for i in range(e_b)):
                        is_price_line=True;c_o=current_row_cells[item_total_price_value_col_idx];current_price_line_info={"value":c_o.value,"coord":c_o.coordinate,"type":"CLASSIC"}
                elif smeta_type is SmetaType.TOTAL_SMALL_CASE and c_val.startswith(KEY_TOTAL_SMALL_CASE):
                    m_jk=check_merge(ws,row_idx,9,10)
                    if m_jk:is_price_line=True;c_o=current_row_cells[item_total_price_value_col_idx];current_price_line_info={"value":c_o.value,"coord":m_jk,"type":"TOTAL_SMALL_CASE","key":KEY_TOTAL_SMALL_CASE}
                elif smeta_type is SmetaType.TOTAL_UPPER_CASE:
                    k = KEY_TOTAL_ALL_UPPER_CASE if c_val.startswith(KEY_TOTAL_ALL_UPPER_CASE) else KEY_TOTAL_WORK_UPPER_CASE if c_val.startswith(KEY_TOTAL_WORK_UPPER_CASE) else None
                    if k:
                        m_jk = check_merge(ws, row_idx, 9, 10)
                        c_o = current_row_cells[item_total_price_value_col_idx]
                        # Исправленный отступ здесь:
                        if m_jk and c_o.value is not None: 
                            found_upper_prices[k] = {"value": c_o.value, "coord": m_jk}

            is_total_key_r=(smeta_type is SmetaType.TOTAL_SMALL_CASE and c_val.startswith(KEY_TOTAL_SMALL_CASE))or \
                            (smeta_type is SmetaType.TOTAL_UPPER_CASE and(c_val.startswith(KEY_TOTAL_ALL_UPPER_CASE)or c_val.startswith(KEY_TOTAL_WORK_UPPER_CASE)))
            if row_type is None and not is_price_line and not is_total_key_r:
                if cell_a and not is_likely_empty(cell_a.value)and cell_a.data_type!="f":
                    item_nature=get_item_id_nature(cell_a.value)
                    if item_nature in {"integer","decimal"}:row_type="item"

            def _apply_price_to_buffer(b_lst,p_k,p_v,p_c,d_lst):
                if not b_lst:return;logger.debug("Прим.цену'%s'(%s)ко %d поз.",p_k,p_v,len(b_lst))
                for itm in b_lst:itm["col_6_value"]=p_v;itm["col_6_coord"]=p_c
                d_lst.extend(b_lst);b_lst.clear()

            is_pot_le_h_r=header_ak_merge_coord and not a_val.startswith(HEADER_SECTION)and not a_val.startswith(HEADER_SUBSECTION)
            if (smeta_type is SmetaType.TOTAL_UPPER_CASE and active_items and found_upper_prices and
                ((row_type=="item"and item_nature=="integer")or row_type in 
                 {"section_header","subsection_header","local_estimate_footer","section_footer","subsection_footer"}or
                 is_pot_le_h_r)):
                k_apl=KEY_TOTAL_ALL_UPPER_CASE if KEY_TOTAL_ALL_UPPER_CASE in found_upper_prices else KEY_TOTAL_WORK_UPPER_CASE if KEY_TOTAL_WORK_UPPER_CASE in found_upper_prices else None
                if k_apl:p_d=found_upper_prices.pop(k_apl);found_upper_prices.clear();_apply_price_to_buffer(active_items,k_apl,p_d["value"],p_d["coord"],processed_rows)
            
            if row_type == "section_header":
                if pending_subsection: processed_rows.append(pending_subsection); pending_subsection = None
                if pending_section: processed_rows.append(pending_section); pending_section = None
                pending_section = {"type":"header","level":"section","start_row":row_idx,"col_1_coord":header_ak_merge_coord,"col_3_value":a_val,"col_6_value":None,"col_6_coord":None}
            elif row_type == "subsection_header":
                if pending_subsection: processed_rows.append(pending_subsection); pending_subsection = None
                pending_subsection = {"type":"header","level":"subsection","start_row":row_idx,"col_1_coord":header_ak_merge_coord,"col_3_value":a_val,"col_6_value":None,"col_6_coord":None}
            
            elif row_type == "local_estimate_footer":
                norm_f_name = _extract_and_normalize_footer_name(a_val)
                logger.debug(f"Стр.{row_idx}: Футер ЛС. Норм.имя футера: '{norm_f_name}' (из '{a_val}')")
                
                matched_header_candidate: Optional[Dict[str, Any]] = None
                found_candidate_idx = -1

                for c_idx in range(len(potential_local_estimate_candidates) - 1, -1, -1):
                    cand = potential_local_estimate_candidates[c_idx]
                    if cand["start_row"] < row_idx: 
                        if norm_f_name and cand["normalized_name"] == norm_f_name:
                            matched_header_candidate = cand; found_candidate_idx = c_idx; break
                        elif not norm_f_name and not cand["normalized_name"]:
                            matched_header_candidate = cand; found_candidate_idx = c_idx; break
                
                if matched_header_candidate and found_candidate_idx != -1:
                    potential_local_estimate_candidates.pop(found_candidate_idx) 
                    if pending_subsection: processed_rows.append(pending_subsection); pending_subsection = None
                    if pending_section: processed_rows.append(pending_section); pending_section = None
                    
                    price_cell_obj = current_row_cells[footer_price_value_col_idx]
                    matched_header_candidate["col_6_value"] = price_cell_obj.value if price_cell_obj else None
                    matched_header_candidate["col_6_coord"] = footer_price_merged_coord
                    matched_header_candidate["level"] = "local_estimate" 
                    processed_rows.append(matched_header_candidate) 
                    logger.info(f"Стр.{row_idx}: Футер '{a_val}' сопоставлен с заголовком ЛС '{matched_header_candidate['col_3_value']}' (стр. {matched_header_candidate['start_row']}).")
                else:
                    logger.warning("Стр.%s: Футер '%s' не нашел подходящего заголовка локальной сметы среди кандидатов.", row_idx, a_val)

            elif row_type == "section_footer": 
                if pending_subsection: processed_rows.append(pending_subsection); pending_subsection = None
                if pending_section: cell_o=current_row_cells[footer_price_value_col_idx];pending_section["col_6_value"]=cell_o.value if cell_o else None;pending_section["col_6_coord"]=footer_price_merged_coord;processed_rows.append(pending_section);pending_section=None
                else: logger.warning("Стр.%s: '%s' найдено, но нет акт.заг.раздела.",row_idx,a_val)
            elif row_type == "subsection_footer": 
                if pending_subsection: cell_o=current_row_cells[footer_price_value_col_idx];pending_subsection["col_6_value"]=cell_o.value if cell_o else None;pending_subsection["col_6_coord"]=footer_price_merged_coord;processed_rows.append(pending_subsection);pending_subsection=None
                else: logger.warning("Стр.%s: '%s' найдено, но нет акт.заг.подраздела.",row_idx,a_val)

            elif is_price_line and active_items: 
                if smeta_type is SmetaType.CLASSIC: _apply_price_to_buffer(active_items, "CLASSIC", current_price_line_info["value"], current_price_line_info["coord"], processed_rows)
                elif smeta_type is SmetaType.TOTAL_SMALL_CASE: _apply_price_to_buffer(active_items, current_price_line_info["key"], current_price_line_info["value"], current_price_line_info["coord"], processed_rows)
            
            elif row_type == "item": 
                skip=False;ind_cell=current_row_cells[material_col_idx] if len(current_row_cells)>material_col_idx else None
                if is_zero(ind_cell.value if ind_cell else None): skipped_zero_price_individual+=1;skip=True;logger.debug(f"Пропуск(0 инд.ц) стр{row_idx},ID:{a_val},кол{material_col_idx+1}")
                if not skip:
                    itm:Dict[str,Any]={"type":"item","start_row":row_idx,"col_6_value":None,"col_6_coord":None}
                    for i in range(min(5,len(current_row_cells))):itm[f"col_{i+1}_value"]=current_row_cells[i].value;itm[f"col_{i+1}_coord"]=current_row_cells[i].coordinate
                    if item_nature=="decimal":itm["col_6_value"]=ind_cell.value if ind_cell else None;itm["col_6_coord"]=ind_cell.coordinate if ind_cell else None;processed_rows.append(itm)
                    elif item_nature=="integer":active_items.append(itm)

        if smeta_type is SmetaType.TOTAL_UPPER_CASE and active_items and found_upper_prices:
            f_key=KEY_TOTAL_ALL_UPPER_CASE if KEY_TOTAL_ALL_UPPER_CASE in found_upper_prices else KEY_TOTAL_WORK_UPPER_CASE if KEY_TOTAL_WORK_UPPER_CASE in found_upper_prices else None
            if f_key:f_p_data=found_upper_prices.pop(f_key);_apply_price_to_buffer(active_items,f_key,f_p_data["value"],f_p_data["coord"],processed_rows)
            found_upper_prices.clear()
        if active_items:
            logger.warning("Тип:%s.Осталось %d поз.б/цены(доб.как есть).",smeta_type.name,len(active_items))
            for itm_ibf in active_items:logger.debug(f"  -Необр.поз:Стр{itm_ibf['start_row']},ID:{itm_ibf.get('col_1_value','N/A')}")
            processed_rows.extend(active_items);active_items.clear()

        if pending_subsection:processed_rows.append(pending_subsection)
        if pending_section:processed_rows.append(pending_section)
        
        if potential_local_estimate_candidates:
            logger.info(f"Осталось {len(potential_local_estimate_candidates)} неиспользованных кандидатов в заголовки ЛС. Они НЕ будут добавлены в результат, так как для них не найден футер.")
            for cand_debug in potential_local_estimate_candidates:
                 logger.debug(f"  - Неиспользованный кандидат ЛС: '{cand_debug['col_3_value']}' (строка {cand_debug['start_row']})")
        
        processed_rows.sort(key=lambda d:d.get("start_row",float("inf")))
        coords_out:List[List[Optional[str]]]=[];
        for entry in processed_rows:
            if entry.get("type")=="item" and is_zero(entry.get("col_6_value")):skipped_zero_price_total+=1;continue
            r_coords=[None]*len(OUTPUT_HEADERS);e_type=entry.get("type");e_level=entry.get("level")
            
            if e_type=="header":
                r_coords[0]=get_start_coord(entry.get("col_1_coord"))
                r_coords[2]=entry.get("col_3_value") 
                r_coords[5]=get_start_coord(entry.get("col_6_coord"))
            elif e_type=="item":
                for i in range(5):r_coords[i]=get_start_coord(entry.get(f"col_{i+1}_coord"))
                r_coords[5]=get_start_coord(entry.get("col_6_coord"))
            
            if any(c is not None for c in r_coords)or(e_type=="header" and entry.get("col_3_value")is not None):
                coords_out.append(r_coords)
        
        logger.info("Обр.'%s'зав.Тип:%s.Проп.поз:инд0ц %s,итог0ц %s.Сформ.стр:%s.",path_obj.name,smeta_type.name,skipped_zero_price_individual,skipped_zero_price_total,len(coords_out))
        return OUTPUT_HEADERS,coords_out
    except FileNotFoundError:logger.error("Файл не найден:%s",path_obj);return None,None
    except Exception:logger.critical("Крит.ошб.при обр.файла'%s'",path_obj,exc_info=True);return None,None
    finally:
        if wb:
            try:wb.close()
            except Exception:logger.warning("Не уд.корр.закр.Excel файл'%s'",path_obj,exc_info=True)